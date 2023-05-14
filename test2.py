import atexit
import datetime
from datetime import datetime
from datetime import timedelta
import time
import tda
import config
from config import *
from tda import auth, client
import json
import pandas as pd
from pandas import json_normalize 
from tda.client import Client
from tda.streaming import StreamClient
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from pandas import ExcelWriter
import matplotlib.pyplot as plt
import numpy as np
print("start")
import asyncio
import requests                             # Telegram
# Frequently Changed Variables####
#stock_name = 'AAPL'
#stock_name = 'SOXL'
stock_name = '$SPX.X'
#stock_name = 'NVDA'
#stock_name = 'AI'
#stock_name = 'SPY'
#stock_name = 'MSFT'
#stock_name = 'COST'
#stock_name = 'QQQ'
#stock_name = 'TSLA'
Time_Frame_1    = 5 # options are 1, 5, 10,15
Time_Frame_2    = 10 # options are 5, 10, 10, 15, 30
instance        = '6R3x:'
ema                     = True # Use EMA or SMA for entry/exits
fast_ema                = 8 #5               # 8
slow_ema                = 34 #13 #21              #13
ema1low                 = 20; 	# for option quotes
ema1high                = 200;	# for option quotes
angle_var               = 0.49
angle_var1              = 0.99
investment      = 10000
in_hand_usd     = investment
data_length     = 2000  # how many values
quantity        = 1000                      # Enter how many contracts
#scale_qty       = round(quantity/2)         # How much scaling In/Out to be done ?
scaling         = False                     # Scaling logic needed ?
timed           = False
plot_var        = False                  #	Decide if you need chart
using_wf        = False                  # Use Williams fractals ?
real_trade              = True # want to send order to TD ?
if timed :
    run_var         = 200                        # How many retries or runs
    L1              = "Timed" # Full Day or Timed
else:
    run_var         = 200                        # How many retries or runs
    L1              = "Full" # Full Day or Timed
## End Variables ##########

print(stock_name, "on ",Time_Frame_1," min and " ,Time_Frame_2," min data" , "instance ",instance)
### Authorization ##########
## Make sure to update Config file with User details
try:
    c = auth.client_from_token_file(config.token_path, config.api_key)
except FileNotFoundError:
    from selenium import webdriver
    with webdriver.Chrome() as driver:
        c = auth.client_from_login_flow(
            driver, config.api_key, config.redirect_uri, config.token_path)
#### End Authorization     #######

# Get Quotes
def quotes(symbol):
    try:
        return client.get_quotes(symbol).json()
    except:
        print("error getting quote")
## end Get Quotes ##     

def excel_write(stock,curr_time,msg,row_number,start):
    write_location = ws.cell(row=row_number,column=2)
    time_location = ws.cell(row=row_number,column=1)
    row_number += 1
    write_location.value = msg
    time_location.value = str(start)+ str(curr_time)
    file_name = str(stock) + "_tr.xlsx"
    wb.save(file_name)
    return row_number 
## end excel write ##
## Telegram        
def telegram_bot_sendtext(bot_message):
    try:
       bot_token = config.telegram_bot_token
       bot_chatID = config.telegram_bot_chatID
       send_text = 'https://api.telegram.org/bot' + bot_token + '/sendMessage?chat_id=' + bot_chatID + '&parse_mode=Markdown&text=' + bot_message
       response = requests.get(send_text)
    except:
       print("Telegram error")
#### Group Message #######################################
def telegram_sendtext_group(bot_message):
    try:
       bot_token = config.telegram_bot_token
       bot_chatID = config.telegram_bot_chatID
       bot_channel_ID = '-1001200290520'
   #bot_channel_ID = "-100585592559"
       send_text = "https://api.telegram.org/bot"+bot_token+"/sendmessage?chat_id="+bot_channel_ID+"&text="+bot_message
       response = requests.get(send_text)
       return response.json()
    except:
       print("Telegram error ")
#### End Telegram msg       ########

# Define ATR function
def get_atr(df, length):
    tr = np.maximum(df['high'] - df['low'], np.abs(df['high'] - df['close'].shift()))
    tr = np.maximum(tr, np.abs(df['low'] - df['close'].shift()))
    atr = np.zeros(len(df))
    atr[length-1] = np.mean(tr[:length])
    for i in range(length, len(df)):
        atr[i] = ((length - 1) * atr[i-1] + tr[i]) / length
    return atr

####### End ATR #######################################
#### Heiken Ashi
def heikin_ashi(df):
    ha_close = (df['open'] + df['high'] + df['low'] + df['close']) / 4
    ha_open = (df['open'].shift(1) + df['close'].shift(1)) / 2
    ha_high = df[['high', 'open', 'close']].max(axis=1)
    ha_low = df[['low', 'open', 'close']].min(axis=1)
    df['close'] = ha_close
    df['open'] = ha_open
    df['high'] = ha_high
    df['low'] = ha_low
    return df[['open', 'high', 'low', 'close','volume','datetime']]

#####End Heiken ashi
#### RSI ###
def rsi(df, periods , ema ):
    close_delta = df['close'].diff()
    # Make two series: one for lower closes and one for higher closes
    up = close_delta.clip(lower=0)
    down = -1 * close_delta.clip(upper=0)    
    if ema == True:
	# Use exponential moving average
        ma_up = up.ewm(com = periods - 1, adjust=True, min_periods = periods).mean()
        ma_down = down.ewm(com = periods - 1, adjust=True, min_periods = periods).mean()
    else:
        # Use simple moving average
        ma_up = up.rolling(window = periods, adjust=False).mean()
        ma_down = down.rolling(window = periods, adjust=False).mean()
        rsi = ma_up / ma_down
    rsi = 100 - (100/(1 + rsi))
    return rsi
#### END RSI
# Define linear regression channel function
def linear_regression_channel(df, length):
    lr = np.polyfit(np.arange(len(df)), df['close'], deg=1)
    lr_upper = lr[0] * np.arange(len(df)) + lr[1] + (2 * np.std(df['close']))
    lr_lower = lr[0] * np.arange(len(df)) + lr[1] - (2 * np.std(df['close']))
    return lr_upper, lr[0] * np.arange(len(df)) + lr[1], lr_lower

def linreg_candles(df):
    linreg_length = 11
    signal_length = 9
    linreg = False
    # Calculate the linear regression values for each price type
    bopen = linreg(df['open'].values, linreg_length) if linreg else df['open'].iloc[-1]
    bhigh = linreg(df['high'].values, linreg_length) if linreg else df['high'].iloc[-1]
    blow = linreg(df['low'].values, linreg_length) if linreg else df['low'].iloc[-1]
    bclose = linreg(df['close'].values, linreg_length) if linreg else df['close'].iloc[-1]
    # Calculate the signal line
    signal = df['close'].iloc[-signal_length:].ewm(span=signal_length).mean().iloc[-1]
    # Generate buy and sell signals based on crossovers of the Linear Regression Candles and the signal line
    buy  = (bclose > signal)
    sell = (bclose < signal)
    # Return the candle colors and signal line
    #return (bopen, bhigh, blow, bclose, signal, buy, sell)
    return (buy, sell)

################################################################

# common Variables##############################################
run                     = 0
prev_HR                 = 0
prev_day                = 0
profit_PCT              = 2
msg                     = ""
qty                     = 1
first_run               = True
x                       = 0    # for messaging
Logic1_buy              = False 	# Buy status of Logic
Logic1_sell             = False 	# Sell status of Logic
Logic1_short            = False 	# short status of Logic
Logic1_buy_cond         = False 	# Buy Condition of Logic
Logic1_sell_cond        = False 	# Sell Condition of Logic
in_hand_Logic1          = 0 	        # How many you have in hand
Logic1_PL               = -0 	        # Actual profit realized
Logic1_max_PL           = -0 	        # Maximum profit realized
Logic1_min_PL           = 0 	        # Minimum/Drawback, profit realized
Logic1_T_Count          = 0 	        # Transaction count
option_buy_value        = 0 	        # Buy Value
option_sell_value       = 0 	        # Sell Value
Logic1_short_value      = 0             # Short Value
Logic1_Dn               = False         # Done for the day no more trades.
Logic1_Net_PL           = 0
loss_Count              = 0
option_Trail_stop       = 0
option_run              = 0             # count how many runs today
total_run               = 0             # count how many runs total
option_entry_limit      = 0
option_SL_hit           = False         # Stop Loss hit 
option_Profit_Booked    = False         # Profit Booked
wait_count              = 0
option_symbol           = ""
bull_wf                 = False
bear_wf                 = False
buy                     = False
sell                    = False
short                   = False
cover                   = False
WF_problem              = False
PL_Doller               = 0
prev_min                = -60
valid_symbol            = True
fast_ema_up             = False
fast_ema_dn             = False
slow_ema_up             = False
slow_ema_dn             = False
scale_in_c              = False
scale_in_c_done         = False
scale_in_p              = False
scale_in_p_done         = False
scale_out_c             = False
scale_out_c_done        = False
scale_out_p             = False
scale_out_p_done        = False
up_angle                = False
dn_angle                = False
up_exit_angle           = False
dn_exit_angle           = False
Pre_WF_bull             = False
Pre_WF_bear             = False
contract_price          = 0
run_days                = 0
option_direction        = "X"
ATM_option              = "xx"
ATM_symbol              = "xx"
Logic1_PLimit           = 0.255 # 0.5 = 50%  How much greedy you are ?
Logic1_PLimit_initial   = 0.01 # 0.5 = 50%  How much greedy you are ?
Logic1_PLimit_increment = 0.01
option_PLimit           = 1.45 # 0.5 = 50%  How much greedy you are ?
option_PLimit_increment = 0.10
Logic1_SLimit_initial   = 0.2 # 0.5 = 50% Default Trail Stop Trail Stop limit after Entry
Logic1_SLimit           = Logic1_SLimit_initial # 0.5 = 50% Default Trail Stop Trail Stop limit after Entry
Logic1_move_stop        = 0.15
WF_bull_count           = 0
WF_bear_count           = 0
profit_Count            =  0
loss_Count              =  0
close                   =  0

value                   = pd.Series([1])
daily_profit            = 0
max_option = 0
option_sell = 0
new_entry = False
up_dir = 0
dn_dir = 0
############################
def make_webdriver():
    # Import selenium here because it's slow to import
    from selenium import webdriver

    driver = webdriver.Chrome()
    atexit.register(lambda: driver.quit())
    return driver
############################
client = tda.auth.easy_client(config.api_key, config.redirect_uri,config.token_path, make_webdriver) # Create a new client
############################
def get_new_data(timeframe): ### get latest data
    extended_hours = True
    # 240 is for day
    if timeframe == 1:
        r           = client.get_price_history_every_minute(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)
    elif timeframe == 5:
        r           = client.get_price_history_every_five_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)
    elif timeframe == 10:
        r           = client.get_price_history_every_ten_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)
    elif timeframe == 15:
        r           = client.get_price_history_every_fifteen_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)
    elif timeframe == 30:
        r           = client.get_price_history_every_thirty_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)
    elif timeframe == 240:
        r           = client.get_price_history_every_day(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)   
    else :
        r           = client.get_price_history_every_ten_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=extended_hours)     
    return r
## End latest data
now                     = datetime.now()    # Update Time
weekday                 = now.weekday()
start_date              = now - timedelta(days=1) # Pull data from past xx days
start                   = now.strftime("%m%d%y")
print(start, L1)
x1_min                  = client.get_price_history_every_minute(stock_name,start_datetime=start_date,need_extended_hours_data=False)
x5_min                  = client.get_price_history_every_five_minutes(stock_name,start_datetime=start_date,need_extended_hours_data=False)
curr_time               = now.strftime("%H%M%S") 		# get current time                    
## pass info to Excel ##
try:
    file_name = str(stock_name) + "_tr.xlsx"
    wb = openpyxl.load_workbook(file_name)
    wb.save(file_name)
except:
    print("Creating New File..")
    wb = Workbook()
ws = wb.active
try:
    row_number = ws.max_row 
    print(row_number ,"rows.")
except:
    print("creating new file")
msg = "Started New Instance at "+ str(start)
row_number = excel_write(stock_name,curr_time,msg,row_number,start)
msg = ""
### Start While Loop 1 ####
 
while 1==1: #Keep Running always
    try:
        now                         = datetime.now()	    # Update Time
        start_date                  = now - timedelta(days=1)   # Pull data from past xx days
        new_day                     = now.strftime("%D")
        if prev_day != new_day : # Update Authorization on new day
            run_days    += 1
            daily_profit = 0
            client = tda.auth.easy_client(
                config.api_key,
                config.redirect_uri,
                config.token_path,
                make_webdriver)
        prev_day                    = new_day
        prev_HR                     = now.strftime("%H")
        curr_time                   = now.strftime("%H%M%S") 		# get current time
        sell_time                   = False # (curr_time > "145800")
        test                        = False
        morning_time                = (curr_time >= "083030" and curr_time <= "093000") and (weekday >=0 and weekday <= 4)
        aftenoon_time               = (curr_time >= "133000" and curr_time <= "145950") and (weekday >=0 and weekday <= 4)
        no_position                 = (option_symbol == "") # and (delete_symbol == "")
        if timed :
            correct_time            = test or morning_time or aftenoon_time or (no_position == False)
        else :
            correct_time            = test or ((curr_time >= "083030") and (curr_time < "150000")) # Set when you want to execute
        scan_time                   = test or ((curr_time >= "083000") and (curr_time < "150015") and (weekday >=0 and weekday <= 4)) # Set when you want to execute
        if option_symbol == "" :    # Clear Option positions when not in position, Housekeeping
            option_position         = ""
            option_symbol           = ""
            option_SL               = 0
            option_run              = 0 
        #### Start While Loop 2 #####
        while scan_time  :
            now = datetime.now()
            curr_time               = now.strftime("%H%M%S") 		# get current time
            curr_min                = now.strftime("%M") 		# get current time
            sell_time               = (curr_time > "145800")
            weekday                 = now.weekday()
            morning_time            = (curr_time >= "083030" and curr_time <= "103000") and (weekday >=0 and weekday <= 4)
            aftenoon_time           = (curr_time >= "133000" and curr_time <= "145950") and (weekday >=0 and weekday <= 4)
            no_position             = (option_symbol == "") 
            if timed :
                correct_time        = test or morning_time or aftenoon_time or (no_position == False)
            else :
                correct_time        = test or ((curr_time >= "083030") and (curr_time < "150000")) 
            scan_time               = test or ((curr_time >= "082900") and (curr_time < "150015") and (weekday >=0 and weekday <= 4)) # Set when you want to execute
            #if curr_min != prev_min :
            #if (option_symbol != "") or abs(int(curr_min)- int(prev_min)) > 1 : #Time_Frame_1 : # in Minutes 1, 3, 5 ,10 are common   
                #print(abs(curr_min - prev_min))
            if 1 == 1 : # get latest data ######################
                one_minute_scan = (int(curr_min) % Time_Frame_1 == 0)*Time_Frame_1!=1 or (curr_min != prev_min)*Time_Frame_1==1
                prev_min = curr_min
########################Time Frame 1 processing                
                if one_minute_scan or first_run:
                    r = get_new_data(Time_Frame_1) # Get Time Frame 1 candles data
                    data_TF1            = r.json()
                    R_TF1_2               = pd.DataFrame.from_dict(data_TF1['candles'])
                    R_TF1_1 = heikin_ashi(R_TF1_2)

                signal_length = 11
                sma_signal = False
                lin_reg = True
                linreg_length = 7#11
                df = R_TF1_1
                df['bclose'] = np.nan_to_num(df['close'].rolling(window=linreg_length).apply(lambda x: np.polyfit(range(linreg_length), x, 1)[0], raw=True))
                df['bopen'] = np.nan_to_num(df['open'].rolling(window=linreg_length).apply(lambda x: np.polyfit(range(linreg_length), x, 1)[0], raw=True))
                df['signal'] = df["bclose"].tail(signal_length).rolling(window=signal_length).mean()

                signal_value        = round(df['signal'].loc[df['signal'].index.max()],5)
                close_value         = round(df['close'].loc[df['close'].index.max()],5)
                bclose_value        = round(df['bclose'].loc[df['bclose'].index.max()],5)
                bopen_value         = round(df['bopen'].loc[df['bopen'].index.max()],2)
                #R_TF1_1.set_index('datetime', inplace=True)
                #date_rng = pd.date_range(start='2022-03-01', end=None, periods=len(R_TF1_1), freq='5T')
                #R_TF1_1.index = pd.DatetimeIndex(date_rng)
                R_TF1_1.set_index(pd.date_range(start='2023-03-01', periods=len(R_TF1_1), freq='5T'), inplace=True)
                R_TF1 = R_TF1_1.resample('3T').agg({'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last', 'volume': 'sum'})
                R_TF1.reset_index(inplace=True)
                time.sleep(0.5)
 ########################Time Frame 2 processing               
                if int(curr_min) % Time_Frame_2 == 0 or first_run:
                    r = get_new_data(Time_Frame_2) # Get Time Frame 2 candles data
                    data_TF2            = r.json()
                    R_TF2_1               = pd.DataFrame.from_dict(data_TF2['candles'])
                    R_TF2 = heikin_ashi(R_TF2_1)
                    df1 = R_TF2_1
                df1['bclose'] = np.nan_to_num(df1['close'].rolling(window=linreg_length).apply(lambda x: np.polyfit(range(linreg_length), x, 1)[0], raw=True))
                df1['bopen'] = np.nan_to_num(df1['open'].rolling(window=linreg_length).apply(lambda x: np.polyfit(range(linreg_length), x, 1)[0], raw=True))
                df1['signal'] = df1["bclose"].tail(signal_length).rolling(window=signal_length).mean()

                signal_value_1        = round(df1['signal'].loc[df1['signal'].index.max()],5)
                close_value_1         = round(df1['close'].loc[df1['close'].index.max()],5)
                bclose_value_1        = round(df1['bclose'].loc[df1['bclose'].index.max()],5)
                bopen_value_1         = round(df1['bopen'].loc[df1['bopen'].index.max()],2)
                R_TF2_1.set_index(pd.date_range(start='2023-03-01', periods=len(R_TF2_1), freq='5T'), inplace=True)
                R_TF2 = R_TF2_1.resample('3T').agg({'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last', 'volume': 'sum'})
                R_TF2.reset_index(inplace=True)
################################################                
                
                atr_var = 14
                ATRMultiplier       = 1.0*morning_time + 1.0*(not morning_time) # Based on morning Time Close
                R_TF1['TR']         = R_TF1['high']- R_TF1['low']
                R_TF1['ATR']        = R_TF1['TR'].rolling(atr_var).mean()

                R_TF1['OHLC4']      = ( R_TF1['close'] + R_TF1['high']+ R_TF1['low']+ R_TF1['close'])/4
                R_TF2['OHLC4']      = ( R_TF2['close'] + R_TF2['high']+ R_TF2['low']+ R_TF2['close'])/4
                
    ## Calculations ##
                var = 100
                last_50 = R_TF1.tail(var)
                last_50_1 = R_TF2.tail(var)
                
                #R_TF1 = R_TF1_.tail(data_length)
                #R_TF2 = R_TF2_.tail(data_length)
                if ema == True:
                    R_TF1[fast_ema]     = last_50['close'].ewm(span=fast_ema, adjust=True).mean()
                    R_TF1[slow_ema]     = last_50['close'].ewm(span=slow_ema, adjust=True).mean()
                    R_TF2[fast_ema]     = R_TF2['close'].ewm(span=fast_ema, adjust=True).mean()
                    R_TF2[slow_ema]     = R_TF2['close'].ewm(span=slow_ema, adjust=True).mean()
                else:
                    R_TF1[fast_ema]     = last_50['close'].rolling(fast_ema).mean()
                    R_TF1[slow_ema]     = last_50['close'].rolling(slow_ema).mean()
                    R_TF2[fast_ema]     = R_TF2['close'].rolling(fast_ema).mean()
                    R_TF2[slow_ema]     = R_TF2['close'].rolling(slow_ema).mean()

                R_TF1['diff']       = R_TF1[fast_ema].diff() # Calculate the difference between consecutive EMA values
                R_TF1['slope']      = R_TF1['diff'] / 1 # Calculate the slope as the ratio of the difference to the time interval between the data points
                R_TF2['diff']       = R_TF2[fast_ema].diff() # Calculate the difference between consecutive EMA values
                R_TF2['slope']      = R_TF2['diff'] / 1 # Calculate the slope as the ratio of the difference to the time interval between the data points
                
                if ema == False:
                    R_TF1['slope_ema1'] = R_TF1['slope'].ewm(span=fast_ema, adjust=True).mean()
                    R_TF1['slope_ema2'] = R_TF1['slope'].ewm(span=slow_ema, adjust=True).mean()
                    R_TF2['slope_ema1'] = R_TF2['slope'].ewm(span=fast_ema, adjust=True).mean()
                    R_TF2['slope_ema2'] = R_TF2['slope'].ewm(span=slow_ema, adjust=True).mean()
                else:
                    R_TF1['slope_ema1'] = R_TF1['slope'].rolling(fast_ema).mean()
                    R_TF1['slope_ema2'] = R_TF1['slope'].rolling(slow_ema).mean()
                    R_TF2['slope_ema1'] = R_TF2['slope'].rolling(fast_ema).mean()
                    R_TF2['slope_ema2'] = R_TF2['slope'].rolling(slow_ema).mean()
                    
                close               = round(R_TF1['close'].loc[R_TF1['close'].index.max()],2)
                fast_ema_value      = round(R_TF1[fast_ema].loc[R_TF1[fast_ema].index.max()],2)
                slow_ema_value      = round(R_TF1[slow_ema].loc[R_TF1[slow_ema].index.max()],2)
                fast_ema_value_1      = round(R_TF2[fast_ema].loc[R_TF2[fast_ema].index.max()],2)
                slow_ema_value_1      = round(R_TF2[slow_ema].loc[R_TF2[slow_ema].index.max()],2)

                last_50 = R_TF1.tail(var)
                fast_slope     = round(last_50['slope_ema1'].loc[last_50['slope_ema1'].index.max()],2) #* slope_multiplier
                slow_slope     = round(last_50['slope_ema2'].loc[last_50['slope_ema2'].index.max()],2) #* slope_multiplier
                slope = fast_slope > slow_slope
                slope_diff = round((fast_slope - slow_slope) *10,2)

                last_50_1 = R_TF2.tail(var)
                fast_slope_1     = round(last_50_1['slope_ema1'].loc[last_50_1['slope_ema1'].index.max()],2) #* slope_multiplier
                slow_slope_1     = round(last_50_1['slope_ema2'].loc[last_50_1['slope_ema2'].index.max()],2) #* slope_multiplier
                slope_1 = fast_slope > slow_slope
                slope_diff_1 = round((fast_slope_1 - slow_slope_1) *10,2)
                
                
                if stock_name == '$SPX.X'  :
                    fast_ema_slope      = round(last_50['slope'].loc[last_50['slope'].index.max()],2)
                    slow_ema_slope      = round(last_50_1['slope'].loc[last_50_1['slope'].index.max()],2)
                else:
                    fast_ema_slope      = round(last_50['slope'].loc[last_50['slope'].index.max()]*10,2)
                    slow_ema_slope      = round(last_50_1['slope'].loc[last_50_1['slope'].index.max()]*10,2)
                close_value     = round(R_TF1['close'].loc[R_TF1['close'].index.max()],5)
                if first_run  :
                    print(stock_name," ",start_date," ", curr_time)
                    fig2 = plt.figure()# For Charting
                    ax1 = fig2.add_subplot(111)
                    k = 0
                
                ATR_value           = (last_50['ATR'].loc[last_50['ATR'].index.max()])
                TR_value            = (last_50['TR'].loc[last_50['TR'].index.max()])
                ATR_high            = (close_value + ATR_value*ATRMultiplier)
                ATR_low             = (close_value - ATR_value*ATRMultiplier)
                standard_dev        = (last_50['OHLC4'].std())
                SD_high             = (close_value + standard_dev*ATRMultiplier)
                SD_low              = (close_value - standard_dev*ATRMultiplier)
                               
                

# Load data from a CSV file (or any other source)
                df = R_TF1 # R_TF1.tail(var)

# Calculate linear regression values
                df['lr_upper'],df['lr'],df['lr_lower']=linear_regression_channel(R_TF1, linreg_length)
                #df['bopen'], df['bhigh'], df['blow'], df['bclose'], df['signal'] = linreg_candles(df)
                
                
                # Calculate price changes and direction (up or down)
                df['price_change']  = df['close'] - df['open']
                df['direction']     = df['price_change'].apply(lambda x: 1 if x > 0 else (-1 if x < 0 else 0))
                direction_value     = (df['direction'].loc[df['direction'].index.max()])
                # Calculate price changes and direction (up or down)
                R_TF2['price_change']  = R_TF2['close'] - R_TF2['open']
                R_TF2['direction']     = R_TF2['price_change'].apply(lambda x: 1 if x > 0 else (-1 if x < 0 else 0))
                direction_value_2     = (R_TF2['direction'].loc[R_TF2['direction'].index.max()])
                
                # Calculate volume at each time interval
                df['volume_change'] = df['volume'] * df['direction']
                # Calculate cumulative volume and OBV
                df['cumulative_volume'] = df['volume_change'].cumsum()
                df['obv']           = df['cumulative_volume']
                ema_period1         = 9
                ema_period2         = 21
                ema_obv = False
                if ema_obv == True:
                    df['obv_ema']       = df['obv'].ewm(span=ema_period1, adjust=False).mean()
                    df['obv_ema2']      = df['obv'].ewm(span=ema_period2, adjust=False).mean()
                else:
                    df['obv_ema']       = df['obv'].rolling(ema_period1).mean()
                    df['obv_ema2']      = df['obv'].rolling(ema_period2).mean()
                #df['obv_above_ema'] = (df['obv'] > df['obv_ema']) & (df['obv_ema'] > df['obv_ema2']) #(df['obv'] > df['obv_ema']) and (df['obv_ema'] > df['obv_ema2'])#.astype(int)
                df['obv_above_ema'] = (df['obv_ema'] > df['obv_ema2']).astype(float)
                #df['obv_below_ema'] = (df['obv'] < df['obv_ema']) & (df['obv_ema'] < df['obv_ema2']) #(df['obv'] < df['obv_ema']).astype(int)
                df['obv_below_ema'] = (df['obv_ema'] < df['obv_ema2']).astype(float)
                above_ema_value     = (df['obv_above_ema'].loc[df['obv_above_ema'].index.max()])
                below_ema_value     = (df['obv_below_ema'].loc[df['obv_below_ema'].index.max()])
                
    ################################################
                #linreg = 7
                #df['buy'],df['sell'] = linreg_candles(R_TF1)
                #print("new way",(R_TF1['buy'].loc[R_TF1['buy'].index.max()]), (R_TF1['sell'].loc[R_TF1['sell'].index.max()]))
                call_1 = (bclose_value_1 > signal_value_1) and (fast_ema_value_1 > slow_ema_value_1 ) #and direction_value == 1 #and above_ema_value == True ##// close  > signal
                put_1  = (bclose_value_1 < signal_value_1) and (fast_ema_value_1 < slow_ema_value_1 ) #and direction_value== -1##and below_ema_value == True #  and (bopen_value < signal_value)
                call = call_1 and (bclose_value > signal_value) and (fast_ema_value > slow_ema_value ) #and direction_value == 1 #and above_ema_value == True ##// close  > signal
                put  = put_1 and (bclose_value < signal_value) and (fast_ema_value < slow_ema_value ) #and direction_value== -1##and below_ema_value == True #  and (bopen_value < signal_value) 
                #print("call", call, "put", put, "bclose_value", bclose_value, "signal_value",signal_value, bclose_value > signal_value )#,"fast_ema_value", fast_ema_value , "slow_ema_value",slow_ema_value)
                #print("slope", slope,"fast_slope",round(fast_slope*10,2),"slow_slope",round(slow_slope*10,2), slope_diff)           
                print('Dir1', direction_value, "Dir2", direction_value_2,"above_OBV" ,above_ema_value,"below_OBV",below_ema_value, "slope_diff", slope_diff,"fast_ema_slope", fast_ema_slope )
                print( "call=",call, bclose_value," >  ",signal_value, bclose_value > signal_value, "AND", fast_ema_value," > ",slow_ema_value , fast_ema_value > slow_ema_value ," and ", slope_diff,slope_diff >= angle_var)
                print( "put=",put, bclose_value," < ",signal_value, bclose_value < signal_value, "AND", fast_ema_value," < ",slow_ema_value , fast_ema_value < slow_ema_value, " and ", slope_diff,-slope_diff <= -angle_var)
                #print( "call=",call,"put=",put,"bclose_value", bclose_value," >/<  signal_value ",signal_value, bclose_value > signal_value, "AND fast_ema_value ", fast_ema_value," >/< slow_ema_value ",slow_ema_value , fast_ema_value > slow_ema_value ," and slope_diff ", slope_diff,slope_diff >= angle_var)
            
                sell            = (put == True)     or (option_sell > 50) or (fast_ema_value < slow_ema_value ) or  (bclose_value < signal_value)#or   slope == False or ((dn_dir >2))) #and (below_ema_value >0))) #and not short # ((call == False) or (put == True)) and not short
                cover           = (call == True)    or (option_sell > 50) or (fast_ema_value > slow_ema_value ) or  (bclose_value > signal_value)# or   slope == True or ((up_dir >2)) #and (above_ema_value >0)) #and not buy   # ((put == False) or (call == True)) and not buy          
                buy             = (call == True)    and (sell == False)  and (slope_diff >= angle_var   or fast_ema_slope >= angle_var1)  and (put == False) and slope ==True #and (up_dir >1) #and (above_ema_value >0)
                short           = (put == True)     and (cover == False) and (slope_diff <= -angle_var or fast_ema_slope <= -angle_var1)and (call == False) and slope == False #and (dn_dir >1) #and (below_ema_value >0)
                
                print(curr_time,close,":",fast_ema_slope,":",slow_ema_slope)#,":",fast_ema_up,":",up_angle)#,":",WF_bull_count,":=:",fast_ema_dn,":", dn_angle,":",WF_bear_count)
               ### END STRATEGY // End Calculations ############################
    # Options Calculations ######################
            if (buy == True or short == True) :#and (option_position == "") :# valid_symbol == False :
                now=datetime.now()
                weekday             = now.weekday() # Monday (0) thru Friday (4)            
                    ## Next Trading day or week day or Valid option day###
                Next_Trd_days       = (1*(weekday>=0)and(weekday<=3))+(3*(weekday==4))+(2*(weekday==5))+(1*(weekday==6)) + 1*(valid_symbol == False)
                Next_date           = now + timedelta(days=Next_Trd_days)
                Next_date           = str(Next_date)
                tomorrow_date       = Next_date[8:10]
                tomorrow_month      = Next_date[5:7]
                tomorrow_year       = Next_date[2:4]                
                    #Next Friday Calculation
                calc_Next_Friday    = (4*(weekday==0))+ (3*(weekday==1))+(2*(weekday==2))+(1*(weekday==3))+(7*(weekday==4))+(6*(weekday==5))+(5*(weekday==6)) + 7*(valid_symbol == False)	
                Next_Friday         = now + timedelta(days=calc_Next_Friday)
                Next_Friday         = str(Next_Friday)
                Friday_date         = Next_Friday[8:10]
                Friday_month        = Next_Friday[5:7]
                Friday_year         = Next_Friday[2:4]
                    ## Today
                Today_date_         = str(now)
                Today_date          = Today_date_[8:10]
                Today_month         = Today_date_[5:7]
                Today_year          = Today_date_[2:4]   
                index = (stock_name == '$SPX.X') #or (stock_name == 'SPY') or (stock_name == 'QQQ')
                #index = False
                if index == True and morning_time:
                    calc_option_date = str(Today_year) +str(Today_month)+str(Today_date)
                    calc_symbol_date = str(Today_month)+str(Today_date) +str(Today_year)
                    sell_time        = (curr_time > "145800")
                if index == True and not morning_time:
                    calc_option_date = str(tomorrow_year) +str(tomorrow_month)+str(tomorrow_date)
                    calc_symbol_date = str(tomorrow_month)+str(tomorrow_date) +str(tomorrow_year)
                    sell_time        = False
                if index == False :
                    calc_option_date = str(Friday_year) +str(Friday_month)+str(Friday_date)
                    calc_symbol_date = str(Friday_month)+str(Friday_date) +str(Friday_year)
                    sell_time        = False
                    #start_date = datetime(year=current_year, month=curr_month, day=current_day)
                    #calc_date = str(current_yr) + str(current_month ) + str(current_day)
                if stock_name == '$SPX.X'  :
                    variable1       = 5
                    option_name     = '.SPXW'
                    symbol_name     = 'SPXW'
                else :
                    variable1 = 1
                    option_name     = "."+ str(stock_name)
                    symbol_name     = stock_name
                if stock_name == 'TSLA' or stock_name == 'NVDA' or stock_name == 'BA':
                    variable1       = 5
                    option_name     = "."+ str(stock_name)
                    symbol_name     = stock_name    
                        
                call_SP = round(close_value)+1          # Latest ITM to reduce Theta decay
                put_SP  = round(close_value)-1
                try:
                    call_SP         = round(ATR_high)       # ATR Levels
                    put_SP          = round(ATR_low)                
                    call_SP         = round(SD_high)        # Standard deviation levels 
                    put_SP          = round(SD_low)
                except:
                    call_SP         = round(close_value)+1
                    put_SP          = round(close_value)-1
                while (call_SP % variable1)  != 0 and ( variable1 == 5): # Needed only for SPX
                    call_SP += 1
                while (put_SP % variable1) != 0 and (variable1 == 5) :  # Needed only for SPX
                    put_SP -= 1
                ATM_call            =  str(option_name) + calc_option_date + "C" + str(call_SP)
                ATM_put             =  str(option_name) + calc_option_date + "P" + str(put_SP)
                ATM_call_symbol     =  str(symbol_name )+ "_" + calc_symbol_date + "C" + str(call_SP)
                ATM_put_symbol      =  str(symbol_name )+ "_" + calc_symbol_date + "P" + str(put_SP)                
                if buy == True : # Time for Call - pass correct parameters
                      ATM_option    =  ATM_call
                      ATM_symbol    =  ATM_call_symbol
                      option_direction = "CALL"
                else:           # Time for Put - pass correct parameters
                     ATM_option     =  ATM_put
                     ATM_symbol     =  ATM_put_symbol
                     option_direction = "PUT"
                result              =  client.get_quotes(ATM_symbol).json()
                valid_symbol        = (result[ATM_symbol]['askPrice'] > 0)
    ## End Options Calculations ###########################
    #########  ###################################         
            Logic1 = "Logic " + str(L1)
            Logic1_active               = (correct_time == True) and (valid_symbol == True)
            
            Logic1_buy_cond             = (buy   and (sell_time == False) and (Logic1_Dn == False) ) and (option_run < run_var) and Logic1_active #and option_SL== False
            print("buy_cond",Logic1_buy_cond   ,"1:", buy               ,"2:",(sell_time == False)  ,"3:", (Logic1_Dn == False)            ,"4:", (option_run < run_var))#,"5:", Logic1_active)
            Logic1_buy_scale_in         = False #(scale_in_c == True) and (scale_in_c_done == False) and (scaling == True)
            Logic1_sell_cond            = ((sell== True)  or  (sell_time == True) ) and Logic1_active
            print("sell_cond",Logic1_sell_cond   ,"1:", put == True     ,"2:",(sell_time == True)   ,"3:", (fast_ema_value < slow_ema_value) ,"4:", (option_sell > 50))#,"5:", Logic1_active)
            Logic1_buy_scale_out        = False #(scale_out_c == True) and (scale_out_c_done == False) and (scaling == True)
            
            Logic1_cover_cond           = ((cover== True) or  (sell_time == True))  and Logic1_active
            Logic1_short_scale_out      = False # (scale_out_p == True) and (scale_out_p_done == False)and (scaling == True)
            Logic1_short_cond           = (short and (sell_time == False) and (Logic1_Dn == False) ) and (option_run < run_var) and Logic1_active #and delete_SL== False
            #Logic1_short_cond           = Logic1_short_cond and option_SL_hit== False and (option_Profit_Booked == False)
            Logic1_short_scale_in       = False # (scale_in_p == True) and (scale_in_p_done == False)and (scaling == True)
            print(curr_time," buy:", Logic1_buy_cond, "sell:",Logic1_sell_cond, "short:",Logic1_short_cond, "cover:",Logic1_cover_cond, "option_sell:",option_sell)#, Logic1_active)

    ###### keep updating In Position Call for Exit    
            if option_symbol != "": ## In Position keep checkiong for Exit
                result = client.get_quotes(option_symbol).json()
                option_ask_price        = result[option_symbol]['askPrice']
                option_bid_price        = result[option_symbol]['bidPrice']
                option_mark_price       = round(((option_bid_price + option_ask_price) / 2),2) 
                print(curr_time, option_position,close,option_buy_value ,option_mark_price, "Tr SL", option_Trail_stop)
                option_Trail_stop       = round(max(max(option_entry_limit,option_mark_price) * (1 - Logic1_SLimit),option_Trail_stop), 1) # TRAIL STOP ##  
                option_tp_price         = round(option_entry_limit * (1 + Logic1_PLimit), 2) # TAKE PROFIT##
                option_move_stop_price  = round(option_entry_limit * (1 + Logic1_move_stop), 2) # Move Trail Stop##
                if (option_mark_price > option_move_stop_price): # Move stops after certain profit
                    Logic1_SLimit = 0.05                
                if (option_mark_price > option_tp_price): # Time to take profit , move stops close 
                    option_Profit_Booked = True
                    Logic1_SLimit       = 0.02
                    print("Time to take profit")
                option_SL = (option_mark_price < option_Trail_stop)*(option_SL+(option_mark_price < option_Trail_stop))
                if (option_SL > 1): # Trail stop loss hit
                    option_SL_hit       = True
                    print("SL Hit")
                Logic1_sell_cond        = Logic1_sell_cond  or (option_SL_hit == True) or (option_sell > 50)#and option_direction == "CALL"
                Logic1_cover_cond       = Logic1_cover_cond or (option_SL_hit == True) or (option_sell > 50)#and option_direction == "PUT"
                if new_entry == False:
                    #print("Length Last Valid Index and value",len(value), value.last_valid_index(),value.loc[value.index.max()])
                    value.loc[value.index.max()] = option_mark_price
                    #print("last value",value.loc[value.index.max()],"mark p",option_mark_price)
                    new_entry = True
                else:
                    value.loc[value.index.max()+1] = option_mark_price
                ema1_low                = value.ewm(span=ema1low, adjust=True).mean()
                ema1_high               = value.ewm(span=ema1high, adjust=True).mean()
                ema1_low_value          = round(ema1_low.loc[ema1_low.index.max()],3)
                ema1_high_value         = round(ema1_high.loc[ema1_high.index.max()],3)
                option_sell_calc             = (ema1_low_value < ema1_high_value)
                if option_sell_calc :
                    option_sell = option_sell + 1
                else:
                    option_sell = 0
                #print(option_mark_price,"low",ema1_low_value,"high",ema1_high_value)
                if option_sell > 1:
                    max_option = max(max_option,option_mark_price)
                    #print(option_mark_price,"may be its time to sell should have sold at", max_option)
                    #option_sell             = False
                if plot_var :
                    plt.tight_layout()
                    ax1.plot(value,'k'    ,label='Actual')
                    ax1.plot(ema1_low,'r' ,label='ema1_low')
                    ax1.plot(ema1_high,'b',label='ema1_high')
                    plt.pause(0.5)
                    if k <= 1:
                      #plt.title(str(stock))
                      plt.tight_layout()
                      plt.legend(loc="best")
                      k = 10
            
############# SELL Option            
            sell_call_cond =  (Logic1_sell_cond  == True) and Logic1_buy   == True    
            sell_put_cond  =  (Logic1_cover_cond == True) and Logic1_short == True     
            if ((sell_call_cond == True and option_direction == "CALL") or (sell_put_cond == True and option_direction == "PUT")) and (option_position != "") and correct_time:
                option_exit_limit = option_mark_price
                qty = in_hand_qty
                in_hand_qty = 0
                corder = tda.orders.options.option_sell_to_close_limit(option_position, qty,option_exit_limit)
                close1 = round(close,1)
                if real_trade == True:
                    r= c.place_order(config.account_id, corder)
                msg = str(instance)+str(curr_time)+" Sell "+ str(option_position) +" for "+ str(option_exit_limit)+ " at close "+ str(close1)+ "-----"
                #telegram_bot_sendtext(msg)   # Send over Telegram
                telegram_sendtext_group(msg) # group message
                row_number = excel_write(stock_name,curr_time,msg,row_number,start)# record in Excel
                print(msg)
        ### Housekeeping
                option_sell_value       = max(option_mark_price,0.01)
                in_hand_Logic1          = in_hand_Logic1 - qty
                PL_Now                  = round((option_sell_value/max(option_buy_value,0.01)) -1,4) *100
                PL_Doller               = int((option_sell_value - option_buy_value)*100) + PL_Doller
                Logic1_PL               = Logic1_PL + PL_Now
                
                if (PL_Now) > 0:
                    profit_Count +=  1
                else :
                    loss_Count   +=  1
                daily_profit            =  round(daily_profit + PL_Now)
                if daily_profit >= 500 or daily_profit <= -100 :
                    print("done for the day per target made ", daily_profit)
                Logic1_Net_PL           = round(Logic1_Net_PL + PL_Now)
                Logic1_max_PL           = max(Logic1_max_PL,PL_Doller)
                Logic1_min_PL           = min(Logic1_min_PL,PL_Doller)
                print(Logic1_Net_PL,"% ", PL_Doller, "in hand $$", round(PL_Doller + in_hand_usd,1),"*")                
                if option_direction == "CALL" : # Call position
                    Logic1_buy          = False 
                    Logic1_sell         = True
                if option_direction == "PUT" : # PUT position
                    Logic1_short        = False
                    Logic1_cover        = True
                # Clear position
                option_position         = ""
                option_symbol           = ""
                option_direction        = "xx"
                option_Trail_stop       = 0
                Logic1_SLimit           = Logic1_SLimit_initial
                max_option = 0
                option_sell =0
                new_entry = False
                try:
                    print("clearing option info")
                    for x in value.index:                  
                        if len(value) > 1 :    	
                            value.drop(x, inplace = True)
                            ema1_low.drop(x, inplace = True)
                            ema1_high.drop(x, inplace = True)
                except:
                    print("problem clearing option data")
                ax1.clear() # Clear Plot    
    ################# BUY Option###################################################################################333
            buy_call_cond =  ((Logic1_buy_cond   and Logic1_buy   == False)) 
            buy_put_cond  =  ((Logic1_short_cond and Logic1_short == False )) 
            if ((buy_call_cond == True and option_direction == "CALL") or (buy_put_cond == True and option_direction == "PUT")) and correct_time:
            ### Option related ##
                result = client.get_quotes(ATM_symbol).json()
                option_ask_price = result[ATM_symbol]['askPrice']
                option_bid_price = result[ATM_symbol]['bidPrice']                
                option_mark_price = round(((option_bid_price + option_ask_price) / 2),2)
                in_hand_qty = quantity
                qty = quantity
                option_entry_limit = option_mark_price
                option_position = ATM_option
                option_symbol   = ATM_symbol
                option_SL = 0
                corder = tda.orders.options.option_buy_to_open_limit(option_position, qty,option_entry_limit)
                if real_trade == True:
                    r= c.place_order(config.account_id, corder)
                close1 = round(close,1)
                msg = str(instance)+str(curr_time)+" Buy "+ str(option_position) +" for "+ str(option_entry_limit)+ " at close "+ str(close1) + "+++++"
                #telegram_bot_sendtext(msg)   # Send over Telegram
                telegram_sendtext_group(msg) # group message
                row_number = excel_write(stock_name,curr_time,msg,row_number,start)
                print(msg)
            ## Housekeeping
                option_run += 1
                total_run += 1
                option_buy_value        = (option_mark_price)
                contract_price          = contract_price + (option_buy_value *100)
                in_hand_Logic1          = in_hand_Logic1 + qty                
                Logic1_T_Count          += 1
                                
                if option_direction == "CALL" : # Call position
                    Logic1_buy          = True
                    Logic1_sell         = False
                if option_direction == "PUT" : # PUT position
                    Logic1_short        = True
                    Logic1_cover        = False
                option_SL_hit           = False
                option_Profit_Booked    = False

                
    ### Hourly Report ####        
            curr_HR = now.strftime("%H") # Every Hour
            if curr_HR != prev_HR:
                #msg = "testing telegram"
                msg = "Summary for "+str(stock_name) +" at "+ str(curr_time) +" " + str(L1)+" instance " + str(instance)+ "timeframe"+ str(Time_Frame_1)
                msg = msg + "\n" + "Transaction/s: " + str(total_run) + " profit: " + str(profit_Count)+ " loss: " + str(loss_Count) + " Win%:" + str((profit_Count/max(1,total_run))*100)+"%"
                msg = msg + "\n" + "Profit based on tranctions: " + str(round(Logic1_Net_PL,1)) + "%"
                msg = msg + "\n made " + str(PL_Doller) + "$ Max profit was " + str(round(Logic1_max_PL,1))+"$ Min profit was "+ str(Logic1_min_PL)+"$"            
                #msg = msg + "\n" + "Investment of $ " + str(investment) + " is now: $ " +str(investment + PL_Doller)+ " in "+ str(run_days)+" day/s "
                #msg = msg + "\n" + "Net Profit on Total investment: " + str(round(100*((investment + PL_Doller)/investment-1),1))+ " %"
                telegram_bot_sendtext(msg) # Send over Telegram
                telegram_sendtext_group(msg) # group message
                print(msg)
                prev_HR = curr_HR
    ### End Reporting ####
    ### HOUSEKEEPING ####
            event = (option_SL_hit== True or option_Profit_Booked== True) 
            if event == True:
                #print("Delayed next entry due to Profit Booking or SL")
                wait_count = wait_count*event + event
            if ((option_SL_hit == True or option_Profit_Booked == True) and (wait_count > 2)):
                    option_SL_hit = False
                    option_Profit_Booked = False
                    wait_count = 0
                    if option_run >= run_var :
                        print("Done for the day")                
            if first_run == True :
                print(Logic1)
                first_run = False  # Run the code under first run only once.                
            if msg != "": # Clear msg 
                msg = ""
            run +=1
            delay = ((option_symbol == ""))*5 + 10
            time.sleep(delay)
            x += 1
            if x >20 :
                x = 0
        #### End while Loop 2 #####
        deep_sleep = (curr_time >= "151500" and curr_time <= "073000")*1200 + 10
        x += 1
        time.sleep(deep_sleep)
    except:
        print("ERROR, delayed next run *******************************************************")
        time.sleep(10)
